"""Generate empty Excel templates with the expected headers.

Used by ``GET /api/imports/templates/{kind}``. The output is a small xlsx
file (header row + 1 example row) consumable by Excel / LibreOffice / Numbers.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.modules.imports.parsers import COLUMNS, ImportKind

_EXAMPLES: dict[ImportKind, list[object]] = {
    "students": [
        "Aïssata", "Camara", "F", "2014-09-15",
        "ECP-001", "CM2-A", "Mariama Camara", "+224620000000",
    ],
    "teachers": [
        "Mamadou", "Bah", "M", "1985-03-22",
        "ECP-001", "+224620111111", "Mathématiques", "Maîtrise",
    ],
    "schools": [
        "ECP-002", "École Primaire de Kankan",
        "REG-KAN", "Kankan", "Kankan-Centre", "Kankan",
        "Quartier Centre", "+224622222222", 10.3854, -9.3066,
    ],
}


def render_template(kind: ImportKind) -> bytes:
    if kind not in COLUMNS:
        raise ValueError(f"unknown import kind: {kind}")

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("openpyxl returned a workbook without an active sheet")
    ws.title = kind.capitalize()

    headers = COLUMNS[kind]
    ws.append(headers)
    bold = Font(bold=True)
    blue_fill = PatternFill("solid", fgColor="DCE6F1")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = blue_fill

    # One example row to make the expected format obvious
    ws.append(_EXAMPLES[kind])

    # Sensible widths
    for col in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col
        )
        letter = col[0].column_letter
        ws.column_dimensions[letter].width = min(max_len + 4, 32)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
