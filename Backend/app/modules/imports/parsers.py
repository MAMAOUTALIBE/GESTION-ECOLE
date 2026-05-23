"""Pure parsers — turn an uploaded Excel/CSV into validated rows + errors.

Each ``parse_*_workbook`` function:
    * accepts ``bytes`` (xlsx) or ``str`` (csv)
    * returns ``ParseResult(rows=[ParsedRow], summary=ImportSummary)``
    * NEVER touches the DB — that's the service's job during commit

Errors are collected per-row so the UI can render a "preview" with red rows.
A row with ``valid=False`` will be skipped at commit time.
"""
from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Literal

from openpyxl import load_workbook

from app.shared.enums import Gender


# =============================================================
# DATA SHAPES
# =============================================================
@dataclass(slots=True)
class ParsedRow:
    rowIndex: int  # 1-based to match Excel UI numbering
    valid: bool
    errors: list[str] = field(default_factory=list)
    data: dict[str, Any] = field(default_factory=dict)


@dataclass(slots=True)
class ImportSummary:
    total: int
    valid: int
    invalid: int


@dataclass(slots=True)
class ParseResult:
    rows: list[ParsedRow]
    summary: ImportSummary
    headers: list[str]


ImportKind = Literal["students", "teachers", "schools"]

# Expected column order for each kind (case-insensitive header matching).
# These are also used by the templates module to generate empty workbooks.
COLUMNS: dict[ImportKind, list[str]] = {
    "students": [
        "firstName", "lastName", "gender", "birthDate",
        "schoolCode", "classRoomName",
        "guardianName", "guardianPhone",
    ],
    "teachers": [
        "firstName", "lastName", "gender", "birthDate",
        "schoolCode", "phone", "subject", "diploma",
    ],
    "schools": [
        "code", "name", "regionCode", "prefecture", "subPrefecture",
        "commune", "address", "phone", "latitude", "longitude",
    ],
}

REQUIRED: dict[ImportKind, list[str]] = {
    "students": ["firstName", "lastName", "gender", "schoolCode"],
    "teachers": ["firstName", "lastName", "gender", "schoolCode"],
    "schools": ["code", "name", "regionCode"],
}


# =============================================================
# COMMON HELPERS (pure)
# =============================================================
_PHONE_DIGITS = re.compile(r"\D+")


def _clean(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _parse_gender(value: Any) -> Gender | None:
    """Accept FR or EN tokens: M, MALE, GARÇON, GARCON → MALE; F, FEMALE, FILLE → FEMALE."""
    raw = _clean(value)
    if raw is None:
        return None
    upper = raw.upper().replace("Ç", "C")
    if upper in ("M", "H", "MALE", "GARCON", "GARÇON", "MASCULIN"):
        return Gender.MALE
    if upper in ("F", "FEMALE", "FILLE", "FEMME", "FEMININ", "FÉMININ"):
        return Gender.FEMALE
    if upper in ("AUTRE", "OTHER", "X"):
        return Gender.OTHER
    return None


def _parse_date(value: Any) -> date | None:
    """Accept ISO 'YYYY-MM-DD', FR 'DD/MM/YYYY', or a real date object."""
    if value is None:
        return None
    if isinstance(value, date):
        return value
    s = _clean(value)
    if s is None:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return date.fromisoformat(s) if fmt == "%Y-%m-%d" else (
                __import__("datetime").datetime.strptime(s, fmt).date()
            )
        except ValueError:
            continue
    return None


def _parse_phone(value: Any) -> str | None:
    """Strip non-digit characters; require at least 7 digits or return None."""
    s = _clean(value)
    if s is None:
        return None
    digits = _PHONE_DIGITS.sub("", s)
    if len(digits) < 7:
        return None
    # Keep a leading + if it was there
    if s.lstrip().startswith("+"):
        return f"+{digits}"
    return digits


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


# =============================================================
# WORKBOOK / CSV READERS — return list of dict[str, Any]
# =============================================================
def _read_xlsx(file_bytes: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb.active
    if sheet is None:
        return [], []
    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    out: list[dict[str, Any]] = []
    for raw in rows:
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in raw):
            continue  # blank line — skip
        out.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    return headers, out


def _read_csv(text: str) -> tuple[list[str], list[dict[str, Any]]]:
    reader = csv.reader(io.StringIO(text))
    try:
        header_row = next(reader)
    except StopIteration:
        return [], []
    headers = [h.strip() for h in header_row]
    out: list[dict[str, Any]] = []
    for raw in reader:
        if all(not (c or "").strip() for c in raw):
            continue
        out.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    return headers, out


def _read(content: bytes | str) -> tuple[list[str], list[dict[str, Any]]]:
    if isinstance(content, str):
        return _read_csv(content)
    if content[:4] == b"PK\x03\x04":  # xlsx zip magic
        return _read_xlsx(content)
    # Fallback: assume utf-8 csv
    return _read_csv(content.decode("utf-8-sig", errors="replace"))


def _check_required(headers: list[str], kind: ImportKind) -> list[str]:
    """Return the required headers MISSING from the file."""
    expected = {h.lower() for h in COLUMNS[kind]}
    actual = {h.lower() for h in headers if h}
    return [h for h in REQUIRED[kind] if h.lower() in expected and h.lower() not in actual]


def _normalize_row(raw: dict[str, Any]) -> dict[str, Any]:
    """Lower-case the keys so header casing in the file doesn't matter."""
    return {k.lower(): v for k, v in raw.items() if k}


# =============================================================
# PER-KIND VALIDATORS (pure functions — easy to unit test)
# =============================================================
def _validate_student(raw: dict[str, Any], row_index: int) -> ParsedRow:
    norm = _normalize_row(raw)
    errors: list[str] = []

    first = _clean(norm.get("firstname"))
    last = _clean(norm.get("lastname"))
    if not first or len(first) < 2:
        errors.append("firstName: longueur minimale 2")
    if not last or len(last) < 2:
        errors.append("lastName: longueur minimale 2")

    gender = _parse_gender(norm.get("gender"))
    if gender is None:
        errors.append("gender: valeur attendue M/F/AUTRE (ou MALE/FEMALE/OTHER)")

    school_code = _clean(norm.get("schoolcode"))
    if not school_code:
        errors.append("schoolCode: requis")

    birth_raw = norm.get("birthdate")
    birth = _parse_date(birth_raw) if birth_raw not in (None, "") else None
    if birth_raw not in (None, "") and birth is None:
        errors.append("birthDate: format invalide (attendu YYYY-MM-DD ou DD/MM/YYYY)")

    return ParsedRow(
        rowIndex=row_index,
        valid=not errors,
        errors=errors,
        data={
            "firstName": first or "",
            "lastName": last or "",
            "gender": gender.value if gender else None,
            "birthDate": birth.isoformat() if birth else None,
            "schoolCode": school_code or "",
            "classRoomName": _clean(norm.get("classroomname")),
            "guardianName": _clean(norm.get("guardianname")),
            "guardianPhone": _parse_phone(norm.get("guardianphone")),
        },
    )


def _validate_teacher(raw: dict[str, Any], row_index: int) -> ParsedRow:
    norm = _normalize_row(raw)
    errors: list[str] = []

    first = _clean(norm.get("firstname"))
    last = _clean(norm.get("lastname"))
    if not first or len(first) < 2:
        errors.append("firstName: longueur minimale 2")
    if not last or len(last) < 2:
        errors.append("lastName: longueur minimale 2")

    gender = _parse_gender(norm.get("gender"))
    if gender is None:
        errors.append("gender: valeur attendue M/F/AUTRE (ou MALE/FEMALE/OTHER)")

    school_code = _clean(norm.get("schoolcode"))
    if not school_code:
        errors.append("schoolCode: requis")

    birth_raw = norm.get("birthdate")
    birth = _parse_date(birth_raw) if birth_raw not in (None, "") else None
    if birth_raw not in (None, "") and birth is None:
        errors.append("birthDate: format invalide (attendu YYYY-MM-DD ou DD/MM/YYYY)")

    return ParsedRow(
        rowIndex=row_index,
        valid=not errors,
        errors=errors,
        data={
            "firstName": first or "",
            "lastName": last or "",
            "gender": gender.value if gender else None,
            "birthDate": birth.isoformat() if birth else None,
            "schoolCode": school_code or "",
            "phone": _parse_phone(norm.get("phone")),
            "subject": _clean(norm.get("subject")),
            "diploma": _clean(norm.get("diploma")),
        },
    )


def _validate_school(raw: dict[str, Any], row_index: int) -> ParsedRow:
    norm = _normalize_row(raw)
    errors: list[str] = []

    code = _clean(norm.get("code"))
    name = _clean(norm.get("name"))
    region_code = _clean(norm.get("regioncode"))
    if not code:
        errors.append("code: requis")
    if not name:
        errors.append("name: requis")
    if not region_code:
        errors.append("regionCode: requis")

    lat = _parse_float(norm.get("latitude"))
    lng = _parse_float(norm.get("longitude"))
    raw_lat = norm.get("latitude")
    raw_lng = norm.get("longitude")
    if raw_lat not in (None, "") and lat is None:
        errors.append("latitude: doit être un nombre")
    if raw_lng not in (None, "") and lng is None:
        errors.append("longitude: doit être un nombre")
    if lat is not None and not -90 <= lat <= 90:
        errors.append("latitude: hors plage (-90, 90)")
    if lng is not None and not -180 <= lng <= 180:
        errors.append("longitude: hors plage (-180, 180)")

    return ParsedRow(
        rowIndex=row_index,
        valid=not errors,
        errors=errors,
        data={
            "code": code or "",
            "name": name or "",
            "regionCode": region_code or "",
            "prefecture": _clean(norm.get("prefecture")),
            "subPrefecture": _clean(norm.get("subprefecture")),
            "commune": _clean(norm.get("commune")),
            "address": _clean(norm.get("address")),
            "phone": _parse_phone(norm.get("phone")),
            "latitude": lat,
            "longitude": lng,
        },
    )


_VALIDATORS = {
    "students": _validate_student,
    "teachers": _validate_teacher,
    "schools": _validate_school,
}


# =============================================================
# PUBLIC ENTRY POINT
# =============================================================
def parse_workbook(content: bytes | str, kind: ImportKind) -> ParseResult:
    """Parse an uploaded Excel/CSV for ``kind`` and validate every row.

    Returns a ``ParseResult`` with per-row errors so the UI can render a
    preview before commit. The DB is never touched here.
    """
    if kind not in _VALIDATORS:
        raise ValueError(f"unknown import kind: {kind}")

    headers, raw_rows = _read(content)
    missing = _check_required(headers, kind)
    if missing:
        # Whole file is unusable — return a single header-level error
        return ParseResult(
            rows=[
                ParsedRow(
                    rowIndex=1,
                    valid=False,
                    errors=[f"En-têtes obligatoires manquants : {', '.join(missing)}"],
                )
            ],
            summary=ImportSummary(total=0, valid=0, invalid=1),
            headers=headers,
        )

    validator = _VALIDATORS[kind]
    parsed = [validator(raw, i + 2) for i, raw in enumerate(raw_rows)]  # +2: header is row 1
    valid_count = sum(1 for r in parsed if r.valid)
    return ParseResult(
        rows=parsed,
        summary=ImportSummary(
            total=len(parsed), valid=valid_count, invalid=len(parsed) - valid_count
        ),
        headers=headers,
    )
