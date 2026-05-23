"""Phase 7 contract tests — Library + Imports.

Pydantic validation, OpenAPI surface, and pure-function parsers/templates.
DB-bound paths (LibraryService.list_inventory, ImportsService.commit dispatch
to Celery) live in tests/integration/ in a later phase.
"""
import csv
import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook
from pydantic import ValidationError

from app.modules.imports.parsers import (
    COLUMNS,
    parse_workbook,
)
from app.modules.imports.schemas import (
    ImportCommitRequest,
    ImportPreviewRow,
)
from app.modules.imports.templates import render_template
from app.modules.library.schemas import (
    LibraryInventoryQuery,
    LibraryLoansQuery,
)
from app.shared.enums import LibraryLoanStatus, LibraryStockStatus


# =====================================================================
# OpenAPI: every Phase 7 endpoint must be discoverable
# =====================================================================
@pytest.mark.asyncio
async def test_openapi_exposes_phase7_endpoints(async_client: AsyncClient) -> None:
    response = await async_client.get("/openapi.json")
    assert response.status_code == 200
    paths = response.json()["paths"]

    for url in (
        # Library (NestJS contract)
        "/api/library/inventory",
        "/api/library/loans",
        # Imports (greenfield)
        "/api/imports/templates/{kind}",
        "/api/imports/{kind}/preview",
        "/api/imports/{kind}/commit",
    ):
        assert url in paths, f"Missing endpoint: {url}"


# =====================================================================
# Library — query DTO defaults & validation
# =====================================================================
def test_library_inventory_query_defaults() -> None:
    q = LibraryInventoryQuery()
    assert q.page == 1
    assert q.pageSize == 100
    assert q.status is None


def test_library_inventory_query_strips_whitespace() -> None:
    q = LibraryInventoryQuery(search="  cahier  ")
    assert q.search == "cahier"


def test_library_inventory_query_accepts_status_enum() -> None:
    q = LibraryInventoryQuery(status=LibraryStockStatus.SHORTAGE)
    assert q.status == LibraryStockStatus.SHORTAGE


def test_library_loans_query_defaults() -> None:
    q = LibraryLoansQuery()
    assert q.page == 1 and q.pageSize == 100 and q.status is None


def test_library_loans_query_status_enum() -> None:
    q = LibraryLoansQuery(status=LibraryLoanStatus.LATE)
    assert q.status == LibraryLoanStatus.LATE


# =====================================================================
# Imports — parser pure functions
# =====================================================================
def _make_xlsx(headers: list[str], rows: list[list[object]]) -> bytes:
    """Build a small in-memory xlsx from headers + rows."""
    from openpyxl import Workbook  # local import keeps test setup obvious

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(headers: list[str], rows: list[list[object]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for r in rows:
        writer.writerow(r)
    return buf.getvalue().encode("utf-8")


def test_parse_workbook_unknown_kind_raises() -> None:
    with pytest.raises(ValueError):
        parse_workbook(b"PK\x03\x04dummy", "unknown")  # type: ignore[arg-type]


def test_parse_students_xlsx_happy_path() -> None:
    xlsx = _make_xlsx(
        COLUMNS["students"],
        [
            ["Aïssata", "Camara", "F", "2014-09-15",
             "ECP-001", "CM2-A", "Mariama", "+224620000000"],
        ],
    )
    result = parse_workbook(xlsx, "students")
    assert result.summary.total == 1
    assert result.summary.valid == 1
    assert result.summary.invalid == 0
    row = result.rows[0]
    assert row.valid is True
    assert row.errors == []
    assert row.data["firstName"] == "Aïssata"
    assert row.data["gender"] == "FEMALE"
    assert row.data["birthDate"] == "2014-09-15"


def test_parse_students_accepts_french_gender_tokens() -> None:
    xlsx = _make_xlsx(
        COLUMNS["students"],
        [
            ["Mamadou", "Bah", "GARCON", None, "ECP-001", None, None, None],
            ["Awa", "Diallo", "FILLE", None, "ECP-001", None, None, None],
        ],
    )
    result = parse_workbook(xlsx, "students")
    assert result.summary.valid == 2
    assert result.rows[0].data["gender"] == "MALE"
    assert result.rows[1].data["gender"] == "FEMALE"


def test_parse_students_collects_per_row_errors() -> None:
    xlsx = _make_xlsx(
        COLUMNS["students"],
        [
            ["A", "B", "ZZ", "not-a-date", "", None, None, None],
        ],
    )
    result = parse_workbook(xlsx, "students")
    row = result.rows[0]
    assert row.valid is False
    assert any("firstName" in e for e in row.errors)
    assert any("lastName" in e for e in row.errors)
    assert any("gender" in e for e in row.errors)
    assert any("schoolCode" in e for e in row.errors)
    assert any("birthDate" in e for e in row.errors)


def test_parse_students_csv_works_too() -> None:
    csv_bytes = _make_csv(
        COLUMNS["students"],
        [["Fatou", "Sylla", "F", "2013-01-10", "ECP-001", "", "", ""]],
    )
    result = parse_workbook(csv_bytes, "students")
    assert result.summary.valid == 1


def test_parse_skips_blank_rows() -> None:
    xlsx = _make_xlsx(
        COLUMNS["students"],
        [
            ["Aïssata", "Camara", "F", None, "ECP-001", None, None, None],
            [None, None, None, None, None, None, None, None],  # blank
        ],
    )
    result = parse_workbook(xlsx, "students")
    assert result.summary.total == 1


def test_parse_returns_header_error_when_required_missing() -> None:
    # 'schoolCode' missing → whole file unusable
    xlsx = _make_xlsx(
        ["firstName", "lastName", "gender", "birthDate"],
        [["A", "B", "F", None]],
    )
    result = parse_workbook(xlsx, "students")
    assert result.summary.invalid == 1
    assert result.summary.total == 0
    assert any(
        "schoolCode" in err for err in result.rows[0].errors
    )


def test_parse_phone_normalization() -> None:
    xlsx = _make_xlsx(
        COLUMNS["students"],
        [
            ["Aïssata", "Camara", "F", None, "ECP-001",
             None, "Mariama", "+224 620 00 11 22"],
        ],
    )
    result = parse_workbook(xlsx, "students")
    assert result.rows[0].data["guardianPhone"] == "+224620001122"


def test_parse_phone_too_short_returns_none() -> None:
    xlsx = _make_xlsx(
        COLUMNS["students"],
        [
            ["A", "B", "F", None, "ECP-001", None, "Mariama", "123"],  # < 7 digits
        ],
    )
    result = parse_workbook(xlsx, "students")
    assert result.rows[0].data["guardianPhone"] is None


def test_parse_birth_date_french_format() -> None:
    xlsx = _make_xlsx(
        COLUMNS["students"],
        [
            ["Aïssata", "Camara", "F", "15/09/2014", "ECP-001",
             None, None, None],
        ],
    )
    result = parse_workbook(xlsx, "students")
    assert result.rows[0].data["birthDate"] == "2014-09-15"


def test_parse_teachers_happy_path() -> None:
    xlsx = _make_xlsx(
        COLUMNS["teachers"],
        [
            ["Mamadou", "Bah", "M", "1985-03-22",
             "ECP-001", "+224620111111", "Maths", "Maîtrise"],
        ],
    )
    result = parse_workbook(xlsx, "teachers")
    assert result.summary.valid == 1
    assert result.rows[0].data["subject"] == "Maths"


def test_parse_schools_happy_path() -> None:
    xlsx = _make_xlsx(
        COLUMNS["schools"],
        [
            ["ECP-002", "École Kankan", "REG-KAN",
             "Kankan", "Kankan-Centre", "Kankan",
             "Quartier", "+224620111111", 10.39, -9.31],
        ],
    )
    result = parse_workbook(xlsx, "schools")
    assert result.summary.valid == 1
    row = result.rows[0]
    assert row.data["latitude"] == pytest.approx(10.39)
    assert row.data["longitude"] == pytest.approx(-9.31)


def test_parse_schools_rejects_out_of_range_coords() -> None:
    xlsx = _make_xlsx(
        COLUMNS["schools"],
        [
            ["ECP-X", "École X", "REG-KAN",
             None, None, None, None, None, 200.0, -200.0],
        ],
    )
    result = parse_workbook(xlsx, "schools")
    row = result.rows[0]
    assert row.valid is False
    assert any("latitude" in e for e in row.errors)
    assert any("longitude" in e for e in row.errors)


# =====================================================================
# Templates — render Excel and re-read
# =====================================================================
@pytest.mark.parametrize("kind", ["students", "teachers", "schools"])
def test_render_template_round_trip(kind: str) -> None:
    blob = render_template(kind)  # type: ignore[arg-type]
    assert blob.startswith(b"PK\x03\x04")  # zip magic
    wb = load_workbook(io.BytesIO(blob), read_only=True)
    sheet = wb.active
    assert sheet is not None
    headers = [c.value for c in next(sheet.iter_rows(max_row=1))]
    assert headers == COLUMNS[kind]


def test_render_template_unknown_kind_raises() -> None:
    with pytest.raises(ValueError):
        render_template("teachersx")  # type: ignore[arg-type]


# =====================================================================
# Imports — schemas
# =====================================================================
def test_commit_request_requires_at_least_one_row() -> None:
    with pytest.raises(ValidationError):
        ImportCommitRequest(rows=[])


def test_commit_request_caps_at_10000() -> None:
    with pytest.raises(ValidationError):
        ImportCommitRequest(
            rows=[
                ImportPreviewRow(rowIndex=i, valid=True, errors=[], data={})
                for i in range(10001)
            ]
        )


def test_commit_request_default_skip_invalid_true() -> None:
    dto = ImportCommitRequest(
        rows=[ImportPreviewRow(rowIndex=2, valid=True, errors=[], data={"x": 1})]
    )
    assert dto.skipInvalid is True


# =====================================================================
# Auth-required endpoints
# =====================================================================
@pytest.mark.asyncio
@pytest.mark.parametrize("url", [
    "/api/library/inventory",
    "/api/library/loans",
])
async def test_phase7_library_requires_bearer_token(
    async_client: AsyncClient, url: str
) -> None:
    response = await async_client.get(url)
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_phase7_template_requires_bearer_token(
    async_client: AsyncClient,
) -> None:
    response = await async_client.get("/api/imports/templates/students")
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_phase7_preview_requires_bearer_token(
    async_client: AsyncClient,
) -> None:
    response = await async_client.post(
        "/api/imports/students/preview",
        files={"file": ("data.xlsx", b"PK\x03\x04", "application/octet-stream")},
    )
    assert response.status_code == 401


@pytest.mark.asyncio
async def test_phase7_commit_requires_bearer_token(
    async_client: AsyncClient,
) -> None:
    response = await async_client.post(
        "/api/imports/students/commit",
        json={
            "rows": [
                {"rowIndex": 2, "valid": True, "errors": [], "data": {}}
            ]
        },
    )
    assert response.status_code == 401
